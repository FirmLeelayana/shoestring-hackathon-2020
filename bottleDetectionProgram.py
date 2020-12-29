import cv2
from PIL import Image
from pyzbar.pyzbar import decode
import numpy as np
import pandas as pd
import datetime
import os.path
from os import path
from imageai.Detection.Custom import CustomObjectDetection

detector = CustomObjectDetection()
detector.setModelTypeAsYOLOv3()
detector.setModelPath("C:/Users/Firm/Documents/5. IfM Hackathon [done]/detection_model-ex-039--loss-0008.472.h5") 
detector.setJsonPath("C:/Users/Firm/Documents/5. IfM Hackathon [done]/OD/ShoestringHackathon2020/BottleDetection/SmallTrial/detection_config.json")
detector.loadModel()

def captureOneFrame(prob = 0.5):
    #captures one frame from camera 0, inputs are input_image, output_image paths, it will return list of
    #cropped image vectors (in cv2 format) for bottles only for further analysis, will overwrite the old input image as
    # the new captured frame and overwrite the old analysed image as the new object analysed image
    #no need for any initial pictures to be present, just run and the pictures will automatically be created
    
    cap = cv2.VideoCapture(0) # use 0 if you only have front facing camera
    ret, frame = cap.read() #read one frame
    cv2.imwrite('current_picture.jpg', frame)

    detections = detector.detectObjectsFromImage(input_image="current_picture.jpg", output_image_path="detected_picture.jpg")

    global height_array
    crop_img_list = []
    height_array = []
    
    for detection in detections:

        if detection['name'] == 'bottle' and detection['percentage_probability'] > prob:
            #if we detect it is a bottle, and iterating through all detections
            coords = detection['box_points']
            img = cv2.imread("current_picture.jpg")
            y1 = coords[1]
            y2 = coords[3]
            x1 = coords[0]
            x2 = coords[2]
            crop_img = img[y1:y2, x1:x2]
            crop_img_list.append(crop_img)
            height = y2 - y1
            height_array.append(height)
            
    return crop_img_list

def barcodeReader(image, bgr = (8, 70, 208)):
    #scans image which is in a number array format
    gray_img = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)
    barcodes = decode(gray_img)

    for decodedObject in barcodes:
        points = decodedObject.polygon

        pts = np.array(points, np.int32)
        pts = pts.reshape((-1, 1, 2))
        cv2.polylines(image, [pts], True, (0, 255, 0), 3)

    for bc in barcodes:
        cv2.putText(image, bc.data.decode("utf-8") + " - " + bc.type, (30, 30), cv2.FONT_HERSHEY_SIMPLEX, 1,
                    bgr, 2)

        return "Barcode: {} - Type: {}".format(bc.data.decode("utf-8"), bc.type)

def barcodeReadcv2(cv2image):
    "reads barcode from the cv2 type image specified"
    barcode = barcodeReader(cv2image)
    return str(barcode)

def barcodeScanFromCropped(cropped_array):
    #receives the list of arrays of the cropped cv format images, and iterates through each array (each representing
    # a detected bottle box) and scans for a barcode within it
    #returns array of barcodes responding to each bottle, returns None if no barcode
    
    barcode_array = []

    for array in cropped_array:
        barcode_per_box = barcodeReadcv2(array)

        barcode_array.append(barcode_per_box)
    return barcode_array

def append_df_to_excel(filename, df, sheet_name='Sheet1', startrow=None,
                       truncate_sheet=False, 
                       **to_excel_kwargs):
    """
    Append a DataFrame [df] to existing Excel file [filename]
    into [sheet_name] Sheet.
    If [filename] doesn't exist, then this function will create it.

    Parameters:
      filename : File path or existing ExcelWriter
                 (Example: '/path/to/file.xlsx')
      df : dataframe to save to workbook
      sheet_name : Name of sheet which will contain DataFrame.
                   (default: 'Sheet1')
      startrow : upper left cell row to dump data frame.
                 Per default (startrow=None) calculate the last row
                 in the existing DF and write to the next row...
      truncate_sheet : truncate (remove and recreate) [sheet_name]
                       before writing DataFrame to Excel file
      to_excel_kwargs : arguments which will be passed to `DataFrame.to_excel()`
                        [can be dictionary]

    Returns: None

    (c) [MaxU](https://stackoverflow.com/users/5741205/maxu?tab=profile)
    """
    from openpyxl import load_workbook

    # ignore [engine] parameter if it was passed
    if 'engine' in to_excel_kwargs:
        to_excel_kwargs.pop('engine')

    writer = pd.ExcelWriter(filename, engine='openpyxl')

    # Python 2.x: define [FileNotFoundError] exception if it doesn't exist 
    try:
        FileNotFoundError
    except NameError:
        FileNotFoundError = IOError


    try:
        # try to open an existing workbook
        writer.book = load_workbook(filename)
        
        # get the last row in the existing Excel sheet
        # if it was not specified explicitly
        if startrow is None and sheet_name in writer.book.sheetnames:
            startrow = writer.book[sheet_name].max_row

        # truncate sheet
        if truncate_sheet and sheet_name in writer.book.sheetnames:
            # index of [sheet_name] sheet
            idx = writer.book.sheetnames.index(sheet_name)
            # remove [sheet_name]
            writer.book.remove(writer.book.worksheets[idx])
            # create an empty sheet [sheet_name] using old index
            writer.book.create_sheet(sheet_name, idx)
        
        # copy existing sheets
        writer.sheets = {ws.title:ws for ws in writer.book.worksheets}
    except FileNotFoundError:
        # file does not exist yet, we will create it
        pass

    if startrow is None:
        startrow = 0

    # write out the new sheet
    df.to_excel(writer, sheet_name, startrow=startrow, **to_excel_kwargs)

    # save the workbook
    writer.save()


def checkIfFilePresent(filePath):
    "File path is a string of the path of the file you want to check, using forward slashes"
    return path.exists(filePath)


def writeToSpreadsheet(filePath, barcode, height_diff):
    """ Input is file path of 
    excel spreadsheet via forward slashes, which can exist beforehand or this function will create it. Headers will be
    created if spreadsheet does not exist beforehand, will just append to existing data structure if file already exists.
    """
    dict1 = {}
    
    #Checking if spreadsheet exists yet or not
    filePresent = checkIfFilePresent(filePath)

    if filePresent:
        #This means we will just need to append to the existing dataframe
        headerPresent = False
        datafm = pd.read_excel(filePath)
        index = datafm.index
        number_of_rows = len(index) + 1
    else:
        headerPresent = True
        number_of_rows = 1
        
    ct = datetime.datetime.now() 
    #Index
    dict1['Index'] = [str(number_of_rows)]
    #Barcode ID, will be equal to None if none detected
    dict1['Barcode ID'] = [str(barcode)]
    #Time of scan
    dict1['Time of Scan'] = [str(ct)]
    #Height of object
    dict1['Height of Object'] = [str(height_diff)]
    df = pd.DataFrame(dict1)
    append_df_to_excel(filePath, df, sheet_name='Sheet1', header = headerPresent, index = False)
   
def captureAndWriteToSpreadsheet(excel_file_path):
    '''takes in the location of an excel file path, if doesnt exist it will be created. Takes one image from the camera,
     detects for bottles, draw boxes around it, takes these boxes to check for barcodes, adds all the bottles and 
     their respective barcodes to the excel spreadsheet'''
    cropped_array = captureOneFrame()
    height_diff_array = height_array
    barcode_list = barcodeScanFromCropped(cropped_array)
    
    i = 0
    for barcode in barcode_list:
        writeToSpreadsheet(excel_file_path, barcode, height_diff_array[i])
        i+=1
        
def continuousCapture(excel_file_path, sampling_time = 1000):
    """Continuous loop of function above, with input being excel file path, and sampling time in ms. Break out of loop by 
    pressing q."""
    
    while True:
        captureAndWriteToSpreadsheet(excel_file_path)
        code = cv2.waitKey(sampling_time)
        if code == ord('q'):
            break
            
continuousCapture('C:/Users/Firm/Documents/5. IfM Hackathon [done]/OD/ShoestringHackathon2020/BottleDetection/SmallTrial/bottle_spreadsheet.xlsx', 1000)